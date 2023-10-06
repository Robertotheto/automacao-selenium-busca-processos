from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
import openpyxl
from time import sleep
from dotenv import load_dotenv
from enviar_email import send_email

load_dotenv()

url = 'https://pje-consulta-publica.tjmg.jus.br/'
code_oab = '133864'

driver = webdriver.Chrome()
driver.get(url)
sleep(30)
field_oab = driver.find_element(By.XPATH, '//*[@id="fPP:Decoration:numeroOAB"]')
field_oab.send_keys(code_oab)
dropdown_state = driver.find_element(By.XPATH, '//*[@id="fPP:Decoration:estadoComboOAB"]')
option_state = Select(dropdown_state)
option_state.select_by_visible_text('SP')
button_search = driver.find_element(By.XPATH, '//*[@id="fPP:searchProcessos"]')
button_search.click()
sleep(10)
processos = driver.find_elements(By.XPATH, '//b[@class="btn-block"]')

for processo in processos:
    processo.click()
    sleep(10)
    window = driver.window_handles
    driver.switch_to.window(window[-1])
    driver.set_window_size(1920, 1080)

    number_process = driver.find_elements(By.XPATH, '//*[@id="j_id132:processoTrfViewView:j_id138"]/div/div[2]/div')
    number_process = number_process[0]
    number_process = number_process.text
    sleep(5)
    data_distribution = driver.find_elements(By.XPATH, '//*[@id="j_id132:processoTrfViewView:j_id150"]/div/div[2]')
    data_distribution = data_distribution[0]
    data_distribution = data_distribution.text
    sleep(5)
    movements = driver.find_elements(By.XPATH, '//div[@id="j_id132:processoEventoPanel_body"]//tr[contains(@class,"rich-table-row")]//td//div//div//span')
    list_movements = []
    for movement in movements:
        list_movements.append(movement.text)
    sleep(5)
    workbook = openpyxl.load_workbook('dados.xlsx')

    try:
        # selecionar a planilha
        pagina_processo = workbook[number_process]
        # criar as colunas
        pagina_processo['A1'] = 'Número do Processo'
        pagina_processo['B1'] = 'Data de Distribuição'
        pagina_processo['C1'] = 'Movimentações'
        # inserir os dados
        pagina_processo['A2'] = number_process
        pagina_processo['B2'] = data_distribution
        for index, linha in enumerate(pagina_processo.iter_rows(min_row=2, max_row=len(list_movements), min_col=3, max_col=3)):
            for celula in linha:
                celula.value = list_movements[index]

        workbook.save('dados.xlsx')
        driver.close()
        sleep(5)
        driver.switch_to.window(window[0])
    except Exception as error:
        # criar a planilha
        workbook.create_sheet(number_process)
        # selecionar a planilha
        pagina_processo = workbook[number_process]
        # criar as colunas
        pagina_processo['A1'] = 'Número do Processo'
        pagina_processo['B1'] = 'Data de Distribuição'
        pagina_processo['C1'] = 'Movimentações'
        # inserir os dados
        pagina_processo['A2'] = number_process
        pagina_processo['B2'] = data_distribution
        for index, linha in enumerate(
                pagina_processo.iter_rows(min_row=2, max_row=len(list_movements), min_col=3, max_col=3)):
            for celula in linha:
                celula.value = list_movements[index]

        workbook.save('dados.xlsx')
        driver.close()
        sleep(5)
        driver.switch_to.window(window[0])
# enviar email
driver.quit()


path = "/home/robertofilho/PycharmProjects/automacao-processos-oab/dados.xlsx"
subject = "Relatorio da pesquisa de uma automação"
send_email(subject, path)

